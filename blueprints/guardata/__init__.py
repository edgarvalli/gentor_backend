from flask import Blueprint, send_file, request, session, redirect
from flask.templating import render_template
from openpyxl.workbook.workbook import Workbook
from db import sql_server as db
from functools import wraps
from auth import _build_auth_code_flow
import app_config

guardata = Blueprint("guardata", __name__, url_prefix="/guardata")


def is_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return redirect("/guardata/")
        return f(*args, **kwargs)
    return wrapper


def is_auth_api(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return {"error": True, "message": "Usuario no autenticado"}
        return f(*args, **kwargs)
    return wrapper


@guardata.route("/")
def guardata_index():

    redirect_url = "/guardata/app"

    if not session.get("user"):
        redirect_url = "/guardata/singin"

    return redirect(redirect_url)


@guardata.route("/verify")
@is_auth
def guardata_verify():
    email = session.get("user").get("preferred_username")
    u = db.fetchone(f"select * from Usuarios where email='{email}'")
    if u is None:
        msg = "El usuarios no esta autorizado en usar la App de Importacion Guardata"
        return render_template("guardata/forbidden.html", msg=msg)
    print(u)
    return redirect("/guardata/app")


@guardata.route("/singin")
def guardata_singin():
    # Technically we could use empty list [] as scopes to do just sign in,
    # here we choose to also collect end user consent upfront
    session['app_redirect'] = '/guardata/verify'
    session["flow"] = _build_auth_code_flow(scopes=app_config.SCOPE)
    return redirect(session["flow"]["auth_uri"])


@guardata.route("/app")
@is_auth
def guardata_app():
    return send_file("templates/guardata/app.html")


@guardata.route("/import", methods=['POST'])
@is_auth_api
def import_guardata():
    from openpyxl import load_workbook

    ALLOWED_EXTENSIONS = ['xlsx','xls']
    file = request.files['file']
    ext = file.filename.split('.')
    ext = ext[len(ext) - 1]
    if ext in ALLOWED_EXTENSIONS:
        data = []
        model = request.args.get('model', '')
        wb = load_workbook(file)
        sheet = wb.sheetnames[0]
        sheet = wb[sheet]
        row = 1
        col = 1
        mandatorykey = ['UUID', 'uuid', 'Uuid']
        mkindex = None
        headers = []
        v = sheet.cell(row, col).value

        while v is not None:
            headers.append(v)
            if v in mandatorykey:
                mkindex = col
            col += 1
            v = sheet.cell(row, col).value

        row = 2
        col = 1
        v = sheet.cell(row, mkindex).value

        msg = ''
        sqlerror = False
        while v is not None:
            item = {}
            for h in headers:
                item[h] = sheet.cell(row, col).value
                col += 1

            r = db.insert(model, item)

            if r is None:
                msg += f"{v} guardado correctamente\n"
            else:
                sqlerror = True
                if r.get('code') == '23000':
                    msg += f"El UUID {v} ya existe\n"
                else:
                    msg += f"[{v}]: {r.get('message')}\n"

            row += 1
            col = 1
            v = sheet.cell(row, mkindex).value

        result = {'error': False, 'message': msg}
        if sqlerror:
            result['sqlmessage'] = 'Ocurrieron errores favor de revisar los log'
        else:
            result['sqlmessage'] = 'Se importo todo correcto'
        
        return result

    return {'error': True, 'message': 'Test'}


@guardata.route("/update/cfdis", methods=["POST", "GET"])
@is_auth_api
def update_cfdis():

    data = request.get_json()
    uuids = data.get("data")
    result = {
        "error": False,
        "message": ""
    }

    for uuid in uuids:
        model = data.get("model")
        query = f"UPDATE {model} SET Cancelado=1 WHERE UUID='{uuid}'"
        r = db.commit(query=query)
        if r["error"]:
            result = r
            break

    return result


@guardata.route("/api/fetch/<model>")
@is_auth_api
def fetch(model):
    data = db.fetchall(f"select * from {model}")
    return {
        "error": False,
        "data": data
    }
